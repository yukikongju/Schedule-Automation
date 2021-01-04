#!/usr/bin/python

""" Generate Active Recall Spreadsheet for each Lectures

    1. Get list of lectures for each courses
    2. Generate header for each lectures and give some space in between

"""
import glob
import openpyxl
import os
from Spreadsheet import Spreadsheet

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

GREY = "c1c1c1"
GREEN = ""
col_question_width = 60

class ActiveRecall(Spreadsheet):
    def __init__(self,  *args, **kwargs):
        super(ActiveRecall, self).__init__(*args, **kwargs)

    def generate_spreadsheet(self):
        self.generate_overview_tab()
        self.generate_lectures_tab_for_all_courses()

    def generate_overview_tab(self):
        starting_col = 1
        row_line = 3
        ws = self.wb.active
        ws.title = "Lectures Recall"
        ws.column_dimensions[get_column_letter(starting_col)].width = col_question_width
        for j, course in enumerate(self.courses_lectures_list):
            # get course_name
            course_name = self.courses_names[j]

            # make course header
            row_line += 1 # offset
            ws.cell(column = starting_col, row = row_line).\
                    value = f"{course_name}"
            column_letter = get_column_letter(starting_col)
            ws[f'{column_letter}{row_line}'].fill = PatternFill(
                    fgColor = GREY, fill_type = "solid")
            ws.merge_cells(f'{column_letter}{row_line}:T{row_line}')
            row_line += 2 # offset of 1

            # add all lecture for course
            for i, lecture in enumerate(course):
                ws.cell(column = starting_col, row = row_line).\
                        value = f"{lecture}"
                row_line += 1


    def generate_lectures_tab_for_all_courses(self):
        for j, course in enumerate(self.courses_lectures_list):
            # create the worksheet
            course_name = self.courses_names[j]
            ws = self.wb.create_sheet(course_name)

            # init the variables
            starting_col = 1
            row_line = 4 # starting at line 4
            ws.column_dimensions[get_column_letter(starting_col)].width =\
                    col_question_width
            for i, lecture in enumerate(course):
                ws.cell(column = starting_col, row = row_line).\
                        value = f"{lecture}"
                column_letter = get_column_letter(starting_col)
                ws[f'{column_letter}{row_line}'].fill = PatternFill(
                        fgColor = GREY, fill_type = "solid")
                ws.merge_cells(f'{column_letter}{row_line}:T{row_line}')
                row_line += 12 # offset of 10


    def save_spreadsheet(self):
        print(self.session_name)
        spreadsheet_name = f"Active Review - {self.session_name}.xlsx"
        self.wb.save(spreadsheet_name)

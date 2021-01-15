#!/usr/bin/python

from spreadsheet import Spreadsheet
from constants import Color
from constants import SpreadsheetParameter

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

class ActiveRecall(Spreadsheet):
    def __init__(self,  *args, **kwargs):
        super(ActiveRecall, self).__init__(*args, **kwargs)

    def generate_spreadsheet(self):
        self.generate_overview_tab()
        self.generate_lectures_tab_for_all_courses()

    def generate_overview_tab(self):
        starting_col = 1
        row_line = 3
        ws = self.wb.active
        ws.title = "Lectures Recall"
        ws.column_dimensions[get_column_letter(starting_col)].width =\
                SpreadsheetParameter.QUESTION_COL_WIDTH 
        for j, course in enumerate(self.courses):

            # make course header
            row_line += 1 # offset
            ws.cell(column = starting_col, row = row_line).\
                    value = f"{course.name}"
            column_letter = get_column_letter(starting_col)
            ws[f'{column_letter}{row_line}'].fill = PatternFill(
                    fgColor = Color.GREY, fill_type = "solid")
            ws.merge_cells(f'{column_letter}{row_line}:T{row_line}')
            row_line += 2 # offset of 1

            # add all lecture for course
            for i, lecture in enumerate(course.lectures):
                ws.cell(column = starting_col, row = row_line).\
                        value = f"{lecture}"
                row_line += 1

    def generate_lectures_tab_for_all_courses(self):
        for j, course in enumerate(self.courses):
            # create the worksheet
            course_name = course.name
            ws = self.wb.create_sheet(course.name)

            # init the variables
            starting_col = 1
            row_line = 4 # starting at line 4
            ws.column_dimensions[get_column_letter(starting_col)].width =\
                SpreadsheetParameter.QUESTION_COL_WIDTH 
            for i, lecture in enumerate(course.lectures):
                ws.cell(column = starting_col, row = row_line).\
                        value = f"{lecture}"
                column_letter = get_column_letter(starting_col)
                ws[f'{column_letter}{row_line}'].fill = PatternFill(
                        fgColor = Color.GREY, fill_type = "solid")
                ws.merge_cells(f'{column_letter}{row_line}:T{row_line}')
                row_line += 12 # offset of 10

    def save_spreadsheet(self):
        spreadsheet_name = f"Active Review - {self.session_name}.xlsx"
        self.wb.save(spreadsheet_name)

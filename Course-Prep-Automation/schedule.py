#!/usr/bin/python

# TODO: refractor with new manager parameters

from spreadsheet import Spreadsheet
from constants import WeekParameter
from constants import DaysOfWeek
from constants import Color
from constants import ScheduleParameter
from constants import SpreadsheetParameter

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.rule import Rule
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

class Schedule(Spreadsheet):
    def __init__(self,  *args, **kwargs):
        super(Schedule, self).__init__(*args, **kwargs)

    def save_spreadsheet(self):
        spreadsheet_name = f"Schedule - {self.session_name}.xlsx"
        self.wb.save(spreadsheet_name)

    def generate_spreadsheet(self):
        self.generate_courses_tab()
        self.generate_daily_schedule()
        self.save_spreadsheet()

    def generate_courses_tab(self):
        """ Creating weekly tabs for all courses 
            
            Implementation:
                - From the courses_names and the courses_lectures_list, generate
                  a weekly schedule for all courses by
                        1. Add columns titles
                        2. Schedule lectures by week according to need
                        3. Add conditional formatting for all lectures row

        """
        for j, course in enumerate(self.courses):
            # create the tab with course name
            ws = self.wb.create_sheet(course.name)

            # adding columns titles
            for i, title in enumerate(ScheduleParameter.COL_TITLES):
                column_index = SpreadsheetParameter.STARTING_COL_INDEX_TITLE + i
                ws.cell(column = column_index, row = SpreadsheetParameter.TITLE_ROW).\
                        value = title   # add title
                column_letter = get_column_letter(column_index)
                ws.column_dimensions[f'{column_letter}'].width =\
                      SpreadsheetParameter.TITLE_COL_WIDTH 

            # change lecture column size
            ws.column_dimensions['A'].width = SpreadsheetParameter.LECTURE_COL_WIDTH 
            week_index = 1 # keeping track of the column
            row_index = SpreadsheetParameter.STARTING_ROW # start at gap
            lecture_index = 0
            lecture_col = 1 # column 'A'
            num_tp_per_week = 1

            # separe lecture by weeks
            for lecture in course.lectures:
                # create new week
                if lecture_index % (WeekParameter.NUM_LECTURES_PER_CLASS_PER_WEEK) == 0:
                    row_index += 1
                    # name week
                    ws.cell(column = lecture_col, row = row_index).\
                            value = f"Week {week_index}"
                    # set background color
                    column_letter = get_column_letter(lecture_col)
                    ws[f'{column_letter}{row_index}'].fill = PatternFill(
                            fgColor= Color.GREY, fill_type="solid")
                    # merge row
                    ws.merge_cells(f'{column_letter}{row_index}:T{row_index}')

                    # increment
                    week_index += 1
                    row_index += 1 # skip a line

                # add lecture
                ws.cell(column = lecture_col, row = row_index).\
                        value = lecture

                # add conditional formatting
                yellow_fill = PatternFill(bgColor = Color.YELLOW)
                pending_rule_style = DifferentialStyle(fill = yellow_fill)
                green_fill = PatternFill(bgColor = Color.GREEN)
                completed_rule_style = DifferentialStyle(fill = green_fill)

                for i, _ in enumerate(ScheduleParameter.COL_TITLES):

                    # add pending value to column
                    column_index = SpreadsheetParameter.STARTING_COL_INDEX_TITLE + i
                    ws.cell(column = column_index, row = row_index).\
                            value = "Pending"  
                    column_letter = get_column_letter(column_index)
                    pending_rule = Rule(type = "containsText",
                            operator = "containsText", text = "Pending",
                            dxf = pending_rule_style)
                    pending_rule.formula =\
                        [f'NOT(ISERROR(SEARCH("Pending",{column_letter}{row_index})))']
                    ws.conditional_formatting.add(f'{column_letter}{row_index}',
                            pending_rule)

                    # add conditional formating to make date green
                    completed_rule = Rule(type = "containsText",
                            operator = "containsText", text = "*",
                            dxf = completed_rule_style)
                    completed_rule.formula =\
                        [f'NOT(ISERROR(SEARCH("*",{column_letter}{row_index})))']
                    ws.conditional_formatting.add(f'{column_letter}{row_index}',
                            completed_rule)

                # increment index
                lecture_index += 1
                row_index += 1 # skip a line

                # put the exercices?
                if lecture_index % (WeekParameter.NUM_LECTURES_PER_CLASS_PER_WEEK) == 0:
                    row_index += 1
                    ws.cell(column = lecture_col, row = row_index).\
                            value = f"TP {week_index - 1}"
                    row_index += 1

    def generate_daily_schedule(self):
        """ Generating Daily Schedule from all the course
            
            Notes:
                a. Each lecture has 1 slide preparation and 1 review
                b. Each chapter has one TP associated with it
                c. Some days may be unavailable for learning

            Implementation:
                - Traverse first lecture for all courses before scheduling another
                  one
                - create two temporary list: one for the lectures, one for the
                  slides, and fill the list by popping the lecture form the list

            Parameters:
                - is_weekend_available (bool): True if user can study on weekends

        """
        # setting up schedule parameters
        is_weekend_available = False
        num_tp_per_lecture = 1
        num_slides_preparation_per_lecture = 1
        num_lecture_review_per_lecture = 1

        # create lectures list to watch in order
        ordered_lectures_list = self.get_ordered_lecture_list()

        # create slides for lecture list
        slides_list = ordered_lectures_list.copy()

        # get reference for first worksheet and rename it
        ws = self.wb.worksheets[0]
        ws.title = "Daily Schedule"


        # Create week
        week_count = 1
        row_index = 6 # begin to draw at row 6 
        start_day_column_index = 3 # days of week start at B
        #  while len(ordered_lectures_list) != 0 or len(slides_list) !=0:
        while len(ordered_lectures_list) != 0:
            # create week row
            _week = ws.cell(column = 1, row = row_index, value = f'Week {week_count}')
            ws.merge_cells(start_row = row_index, start_column = 1, end_column
                    = 15, end_row = row_index)
            _week.fill = PatternFill(fgColor = Color.GREY, fill_type= "solid")
            # TODO: ADD date to week
            row_index += 2 # skip a line

            # create week days column
            for i, day in enumerate(DaysOfWeek.DAYS_OF_WEEK_FRENCH):
                column = start_day_column_index + i
                _cell = ws.cell(row = row_index, column = column, value = day)
                ws.column_dimensions[get_column_letter(column)].width =\
                        SpreadsheetParameter.LECTURE_COL_WIDTH
            row_index += 1
            
            # saving row index for backtracking
            start_lecture_index = row_index # to reset row index after each iter
            start_slides_index = start_lecture_index + WeekParameter.NUM_LECTURES_PER_DAY

            # First Column TITLE name
            # Lecture 1 to n
            for i in range(WeekParameter.NUM_LECTURES_PER_DAY):
                ws.cell(column = 1, row = row_index, value = f'Lecture {i+1}')
                row_index += 1
            # Slides 1 to n
            for i in range(WeekParameter.NUM_SLIDES_PER_DAY):
                ws.cell(column = 1, row = row_index, value = f'Slide {i+1}')
                row_index += 1

            # Schedule lectures for the week
            lecture_days_indexes = [1,2,3,4,5] # lectures from monday to friday only
            for col, day in enumerate(lecture_days_indexes):
                for j in range(WeekParameter.NUM_LECTURES_PER_DAY):
                    if len(ordered_lectures_list) != 0 :
                        ws.cell(column = start_day_column_index + col, 
                                row = start_lecture_index + j, 
                                value = ordered_lectures_list.pop(0))
                    else:
                        break 

            # schedule slides for the week
            #  slide_days_indexes = [1,2,3,4,5,7] # samedi is break from slides prep
            slide_days_indexes = [1,2,3,4,5] # slides only on week days
            for col, day in enumerate(slide_days_indexes):
                for j in range(WeekParameter.NUM_SLIDES_PER_DAY):
                    if len(slides_list) != 0 :
                        ws.cell(column = start_day_column_index + col, 
                                row = start_slides_index + j, 
                                value = slides_list.pop(0))
                    else:
                        break 

            # increment variables
            week_count += 1
            row_index += 1

    def get_ordered_lecture_list(self):
        """ Get a list of the order to watch the lecture for all classes
            
            Implementation:
                - Traverse first lecture for all courses before adding the second
                  one
        """
        # initalize the list
        ordered_list = []
        #  retrieve the lectures to watch in order
        lecture_index = 0
        while len(self.courses) != 0:
            # add lectures to list until there are none left
            for i, course in enumerate(self.courses):
                if len(course.lectures) != 0:
                    ordered_list.append(self.courses[i].lectures.pop(0))
                else: # we pop the empty list
                    self.courses.pop(i)
            lecture_index += 1
        return ordered_list


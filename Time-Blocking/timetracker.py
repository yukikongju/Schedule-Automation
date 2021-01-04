#!/usr/bin/python

"""
    0. Create the document
        a. Generate its name generation: Time Tracker - [Month Year]
        b. Generate Sheet for all weeks in the month
        c. Automate the file creation with windows tasks automation

    1. Create a legend 
        a. Properties
            - name (string): name of the tag
            - color (string): color code for the tag
            - id (string): ??? an id if we can't recognize the tag by its color
            - description (string): description of the tag
    2. Generate the template with date and time
    3. Create weekly stats with streamlit

    TODO:
        - [ ] Fix Legend Color
        - [ ] Add Timestamp with 15 minutes increment
        - [ ] Add date above week days

"""
import calendar
import datetime
import openpyxl
import os

from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter

from utils import DAYS_OF_WEEK_FRENCH
from utils import Color

row_date = 2
row_title = row_date + 2 # gap for the row title
column_width = 25

legend_column_index = 11 # column start at index K

# path where the file will be saved
path = "C:/Users/emuli/OneDrive - Universite de Montreal/Bac-Maths-Info/Organization Documents/Time Tracker"

#  template_file = "Time Tracker - Template.xlsx"

legend = {'Active Recall': Color.GREEN, 
          'TPs and Exercices': Color.EMERAUDE,
          'Slides': Color.TURQUOISE,
          #  'Lecture Review': Color.DARK_BLUE,
          'Lecture Review': Color.LIME,
          'Unanswered Questions': Color.BLUE_CLEAR,
          'Unanswered Questions': Color.CYAN,
          'Coding': Color.BLUE_SKY,
          'Management': Color.INDIGO,
          'Training': Color.VIOLET,
          'Chores + Toilettes': Color.ORANGE,
          'Eating': Color.ORANGE,
          'Wasted Time': Color.RED,
          'Social': Color.YELLOW,
          'Sleep': Color.LIGHT_GREY,
          #  'Deplacement': 
          #  'Journaling':
          }

def main():
    """ Create Time Tracker Excel Spreadsheet

        Implementation:
            1. Add legend to the table
            2. Add days of week as columns
            3. Add timestamp as 15 minutes increment
            4. Add date above days of week
            5. Create a stats sheet for weekly-monthly schedule
    
    """
    wb = Workbook() # create workbook
    ws = wb.active

    # find the first monday of the month
    start_date = get_month_first_monday()
    
    # generate week tab
    for i in range(4):
        ws = wb.create_sheet(f"Week {i+1}")
    
        # Add legend to tab
        legend_column_letter = get_column_letter(legend_column_index)
        ws[f'{legend_column_letter}{row_title}'] = "Légende"
        for i, tag in enumerate(legend):
            row_index = row_title + i + 2 # gap of two
            _cell = ws.cell(row = row_index, column = legend_column_index)
            _cell.value = f"{tag}"
            _cell.fill = PatternFill("solid", fgColor= legend[tag][1:])
            _cell.font = Font(color = Color.WHITE[1:])
            
        #  Adjust columm width
        ws.column_dimensions['K'].width = column_width # adjust legend col width

        #  Add days of week
        for i, day in enumerate(DAYS_OF_WEEK_FRENCH):
            column_index = i+2
            _day = ws.cell(column = column_index, row = row_title, value = day)
            _day.alignment = Alignment(horizontal = 'center')
            column_letter = get_column_letter(column_index)
            # adjust days column width
            ws.column_dimensions[f'{column_letter}'].width = column_width 

            # TODO: Add dates above days of week
            _date = ws.cell(column = column_index, row = row_date, value = start_date)
            _date.alignment = Alignment(horizontal = 'center')
            start_date += timedelta(days = 1)

        
        # TODO: Add timestamp
        dummy_year, dummy_month, dummy_day = 2020, 12, 1
        START_DAY = datetime(dummy_year, dummy_month, dummy_day, 
                7,30) # timetable starts at 7h30 AM 
        END_DAY = datetime(dummy_year, dummy_month, dummy_day + 1, 
                1, 30) # timetable ends at 1h30 AM
        INCREMENT = timedelta(minutes = 15)
        timestamp = START_DAY
        num_increment = 1
        ws.column_dimensions['A'].width = 12 # adjust legend col width
        while timestamp < END_DAY:
            row_index = row_title + num_increment
            _timestamp = ws.cell(column = 1, row = row_index, 
                    value = timestamp.strftime("%H:%M %p"))
            timestamp += INCREMENT
            num_increment += 1


    # rename workbook - Month Year

    # save file
    wb.save(os.path.join(path, f'Time Tracker - {get_month_year()}1.xlsx'))

def get_month_first_monday():
    """ Function that finds first monday of the current month
    
        RETURN: (date) 
    """
    this_month = datetime.now().month
    this_year = datetime.now().year
    first_day_of_month = date(this_year, this_month, 1)
    first_monday = first_day_of_month + timedelta(
            days =-first_day_of_month.weekday(), weeks = 1)
    return first_monday


def get_month_year():
    """ Function that retrieve this month date and year
        
        RETURN: (date) December 2020
    """
    this_month = datetime.now().strftime("%B")
    this_year = datetime.now().year
    return f'{this_month} {this_year}'

if __name__ == "__main__":
    main()





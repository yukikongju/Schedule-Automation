#!/usr/bin/python

"""
    1. Create a tab file separating the courses material by week for all files
       specified
            a. Add pending for all columns
                * [slides] - green if lecture preparation has been done
                * [lecture] - green if lecture has been watched
                * [review] - green if lecture review has been done
            b. Separe material by weeks
                * set a number of slides to go through per week
                * set unavailable day
                * create a merged bar for row with week
                * add a start date and end date for the session
    2. Create a daily schedule from the weekly separation of all tabs
    3. Open Lecture Notes for the day automatically

"""

import glob
import openpyxl
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter

#  from . import utils #DAYS_OF_WEEK_FRENCH


wb = Workbook() # init workbook

material_path = "courses_material"

num_slides_per_day = 2 # number of slides to go through per day
num_lectures_per_day = 2 # number of lecture to attend per day
num_exercices_per_day = 1 # number of exercices to do per day

class_lectures_per_week = 2 # num of lecture for a class in a week
class_exercices_per_week = 1

DAYS_OF_WEEK_FRENCH = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi',
        'Dimanche']

LINKS_CELLS = {'Slides Directory':'B1',
        'Notes de Cours One Note': 'B2',
        'Exercices Spreadsheet': 'B3',
        'Table des Matieres': 'B4'
         }

#  title_row = len(LINK_TITLES) + 2
title_row = len(LINKS_CELLS) + 2
start_row = title_row + 2

lecture_col_width = 25
col_title_width = 15

start_index_title = 3 # titles column starts at C

#  GREY = "696969"
GREY = "c1c1c1"
#  PASTEL = "5f9ea0"
#  YELLOW = "fdfd96"
#  YELLOW = "ffcc00"
YELLOW = "ffff33"
#  GREEN = "03c03c"
#  GREEN = "99cc00"
GREEN = "adff2f"

col_titles = ['Slides',             # if lecture has been prepared
              'Lecture',            # if lecture has been attended
              'Review'              # if lecture review has been done
              ]

courses_list = [] # a list of the course name
courses_lectures_list = [] # a list of the lecture list for all courses

def get_courses_material():
    """ Get a list for all courses

        Implementation:
            - Open all files in the material directory and save courses name as
              one list and the lectures for each courses as another one

        Parameters:
            - courses_list = ['course1', 'course2']
            - courses_lectures_list = [['course1_lecture1'],['course2_lecture1']]

    """
    os.chdir(material_path) # change directory
    for text_doc in glob.glob("*.txt"):
        # get course name
        course_name = text_doc.replace(".txt", "")
        # get lecture list
        lectures_list = []
        with open(text_doc) as f:
            lectures_list = [line.strip() for line in f]
        f.close()
        # put data into dataframe
        courses_list.append(course_name) 
        courses_lectures_list.append([lecture for lecture in lectures_list])


def generate_courses_tab():
    """ Creating weekly tabs for all courses 
        
        Implementation:
            - From the courses_list and the courses_lectures_list, generate
              a weekly schedule for all courses by
                    1. Add columns titles
                    2. Schedule lectures by week according to need
                    3. Add conditional formatting for all lectures row

    """
    for j, course_lectures in enumerate(courses_lectures_list):
        # create the tab with course name
        course_name = courses_list[j] # course name is saved at index 0
        ws = wb.create_sheet(course_name)

        # add title to links
        #  for i, title in enumerate(LINK_TITLES):
        #      ws.cell(column = 1, row = i + 1, value = title)

        # adding columns titles
        for i, title in enumerate(col_titles):
            column_index = start_index_title + i
            ws.cell(column = column_index, row = title_row).\
                    value = title   # add title
            column_letter = get_column_letter(column_index)
            ws.column_dimensions[f'{column_letter}'].width =\
                   col_title_width

        # change lecture column size
        ws.column_dimensions['A'].width = lecture_col_width # adjust legend col width

        week_index = 1 # keeping track of the column
        row_index = start_row # start at gap
        lecture_index = 0
        lecture_col = 1 # column 'A'

        # separe lecture by weeks
        for lecture in course_lectures:
            # create new week
            if lecture_index % (class_lectures_per_week) == 0:
                row_index += 1
                # name week
                ws.cell(column = lecture_col, row = row_index).\
                        value = f"Week {week_index}"
                # set background color
                column_letter = get_column_letter(lecture_col)
                ws[f'{column_letter}{row_index}'].fill = PatternFill(
                        fgColor= GREY, fill_type="solid")
                # merge row
                ws.merge_cells(f'{column_letter}{row_index}:T{row_index}')

                # increment
                week_index += 1
                row_index += 1 # skip a line

            # add lecture
            ws.cell(column = lecture_col, row = row_index).\
                    value = lecture

            # add exercices

            yellow_fill = PatternFill(bgColor = YELLOW)
            pending_rule_style = DifferentialStyle(fill = yellow_fill)
            green_fill = PatternFill(bgColor = GREEN)
            completed_rule_style = DifferentialStyle(fill = green_fill)

            for i, _ in enumerate(col_titles):

                # add pending value to column
                column_index = start_index_title + i
                ws.cell(column = column_index, row = row_index).\
                        value = "Pending"  
                column_letter = get_column_letter(column_index)
                pending_rule = Rule(type = "containsText",
                        operator = "containsText", text = "Pending",
                        dxf = pending_rule_style)
                pending_rule.formula =\
                    [f'NOT(ISERROR(SEARCH("Pending",{column_letter}{row_index})))']
                ws.conditional_formatting.add(f'{column_letter}{row_index}',
                        pending_rule)

                # add conditional formating to make date green
                completed_rule = Rule(type = "containsText",
                        operator = "containsText", text = "*",
                        dxf = completed_rule_style)
                completed_rule.formula =\
                    [f'NOT(ISERROR(SEARCH("*",{column_letter}{row_index})))']
                ws.conditional_formatting.add(f'{column_letter}{row_index}',
                        completed_rule)

            # increment index
            lecture_index += 1
            row_index += 1

            # put the exercices?




def generate_daily_schedule():
    """ Generating Daily Schedule from all the course
        
        Notes:
            a. Each lecture has 1 slide preparation and 1 review
            b. Each chapter has one TP associated with it
            c. Some days may be unavailable for learning

        Implementation:
            - Traverse first lecture for all courses before scheduling another
              one
            - create two temporary list: one for the lectures, one for the
              slides, and fill the list by popping the lecture form the list

        Parameters:
            - is_weekend_available (bool): True if user can study on weekends

    """
    # setting up schedule parameters
    is_weekend_available = False
    num_tp_per_lecture = 1
    num_slides_preparation_per_lecture = 1
    num_lecture_review_per_lecture = 1

    # create lectures list to watch in order
    ordered_lectures_list = get_ordered_lecture_list()

    # create slides for lecture list
    slides_list = ordered_lectures_list.copy()

    # get reference for first worksheet and rename it
    ws = wb.worksheets[0]
    ws.title = "Daily Schedule"


    # TODO: Create week
    week_count = 1
    row_index = 6 # begin to draw at row 6 
    start_day_column_index = 3 # days of week start at B
    #  while len(ordered_lectures_list) != 0 or len(slides_list) !=0:
    while len(ordered_lectures_list) != 0:
        # create week row
        _week = ws.cell(column = 1, row = row_index, value = f'Week {week_count}')
        ws.merge_cells(start_row = row_index, start_column = 1, end_column
                = 15, end_row = row_index)
        _week.fill = PatternFill(fgColor = GREY, fill_type= "solid")
        # TODO: ADD date to week
        row_index += 2 # skip a line

        # create week days column
        for i, day in enumerate(DAYS_OF_WEEK_FRENCH):
            column = start_day_column_index + i
            _cell = ws.cell(row = row_index, column = column, value = day)
            ws.column_dimensions[get_column_letter(column)].width =\
                    lecture_col_width
        row_index += 1
        
        # saving row index for backtracking
        start_lecture_index = row_index # to reset row index after each iter
        start_slides_index = start_lecture_index + num_lectures_per_day

        # TODO: First Column TITLE name
        # Lecture 1 to n
        for i in range(num_lectures_per_day):
            ws.cell(column = 1, row = row_index, value = f'Lecture {i+1}')
            row_index += 1
        # Slides 1 to n
        for i in range(num_slides_per_day):
            ws.cell(column = 1, row = row_index, value = f'Slide {i+1}')
            row_index += 1


        # TODO: Schedule lectures for the week
        lecture_days_indexes = [1,2,3,4,5] # lectures from monday to friday only
        for col, day in enumerate(lecture_days_indexes):
            for j in range(num_lectures_per_day):
                if len(ordered_lectures_list) != 0 :
                    ws.cell(column = start_day_column_index + col, 
                            row = start_lecture_index + j, 
                            value = ordered_lectures_list.pop(0))
                else:
                    break 

        # TODO: schedule slides for the week
        #  slide_days_indexes = [1,2,3,4,5,7] # samedi is break from slides prep
        slide_days_indexes = [1,2,3,4,5] # slides only on week days
        for col, day in enumerate(slide_days_indexes):
            for j in range(num_slides_per_day):
                if len(slides_list) != 0 :
                    ws.cell(column = start_day_column_index + col, 
                            row = start_slides_index + j, 
                            value = slides_list.pop(0))
                else:
                    break 


        # increment variables
        week_count += 1
        row_index += 1


    pass

def get_ordered_lecture_list():
    """ Get a list of the order to watch the lecture for all classes
        
        Implementation:
            - Traverse first lecture for all courses before adding the second
              one
    """
    # initalize the list
    ordered_list = []

    # TODO: retrieve the lectures to watch in order
    #  for course_index, course in enumerate(courses_list):
    lecture_index = 0
    while len(courses_lectures_list) != 0:
        # add lectures to list until there are none left
        for i, course in enumerate(courses_lectures_list):
            if len(course) != 0:
            #  if len(courses_lectures_list[i]) != 0:
                #  ordered_list.append(course[lecture_index])
                #  print(courses_lectures_list[i][0])
                ordered_list.append(courses_lectures_list[i].pop(0))
                #  print(course)
            else: # we pop the empty list
                #  print(i)
                #  courses_list.pop(i)
                courses_lectures_list.pop(i)
        lecture_index += 1
    
    #  for lecture in ordered_list:
    #      print(lecture)

    return ordered_list



if __name__ == "__main__":
    #  wb = Workbook() # init workbook
    #  ws = ws.active # get reference for first worksheet
    
    # get courses content from directory
    get_courses_material()

    # generate weekly tabs for all courses in the directory
    generate_courses_tab()
    
    # generate daily schedule
    generate_daily_schedule()

    # TODO: change file path and name
    #  wb.save(os.path.join(path, 'Schedule - Test.xlsx'))
    spreadsheet_name = "Schedule - Session 2.xlsx"
    #  wb.save('Schedule - Test.xlsx')
    wb.save(spreadsheet_name)
